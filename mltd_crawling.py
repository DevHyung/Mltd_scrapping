# -*- encoding:utf8 -*-
import os
import time
from selenium import webdriver
from bs4 import BeautifulSoup
from urllib.request import urlretrieve
from openpyxl import load_workbook, Workbook


def BRANDS(text, num, idx):
    if not os.path.isdir("0" + idx + " " + text):
        os.mkdir("0" + idx + " " + text)

    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="공급사 상품명")  # 공급사 상품명
    ws.cell(row=1, column=2, value="상품명")  # 상품명
    ws.cell(row=1, column=3, value="옵션입력")  # 옵션입력
    ws.cell(row=1, column=4, value="기타")  # 기타
    ws.cell(row=1, column=5, value="공급가")  # 공급가
    ws.cell(row=1, column=6, value="변동가")  # 변동가

    wb.save("0" + idx + " " + text + "/#" + text + ".xlsx")

    bs4 = BeautifulSoup(driver.page_source, 'html.parser')
    Url = bs4.findAll('a', class_='col-md-12 col-sm-12 col-xs-12 nopadding font-16')

    rowNum = 2
    imgNum = 1
    for i in range(0, len(Url)):
        while True:
            answer = input(text + '-' + Url[i].text + "의 데이터를 추출하시겠습니까? y/n : ")
            if answer == "y":
                break
            elif answer == "n":
                break
            else:
                continue

        if answer == 'n':
            continue

        elif answer == 'y':
            driver.get("http://www.mltd.com/" + Url[i]['href'])

            try:
                driver.find_element_by_xpath('//*[@id="sortby"]/option[2]').click()
            except:
                driver.find_element_by_xpath('//*[@id="fysSort"]/option[3]').click()

            while True:
                bs4 = BeautifulSoup(driver.page_source, "html.parser")
                List = bs4.findAll('div', {'class': "item active"})

                for k in range(0, len(List)):
                    url = "https://www.mltd.com" + List[k].find('a')['href']
                    driver.get(url)

                    # 공급사 상품명
                    temp = url.partition('pr-')[2]
                    elem1 = temp.partition('.')[0]

                    # 상품명
                    a = driver.find_element_by_xpath('//*[@id="addprdct"]/p').text
                    elem2 = a.replace('\n', ' / ')

                    # 옵션입력, 기타
                    elem3 = ''
                    elem4 = ''
                    m = 1
                    while True:
                        try:
                            b = driver.find_element_by_xpath(
                                '//*[@id="addprdct"]/div[7]/table/tbody/tr[' + str(m) + ']/td[1]').text
                            if m == 1:
                                elem3 = b
                            else:
                                elem3 = elem3 + ' || ' + b
                            try:
                                btemp = b.partition(' / ')[0]
                            except:
                                btemp = b
                            c = driver.find_element_by_xpath(
                                '//*[@id="addprdct"]/div[7]/table/tbody/tr[' + str(m) + ']/td[2]').text
                            ctemp = c.split(' ')[0]
                            if ctemp == '':
                                ctemp = '100'
                            if m == 1:
                                elem4 += btemp + '-' + ctemp
                            else:
                                elem4 += '/' + btemp + '-' + ctemp
                            m += 1
                        except:
                            break

                    # 공급가 , 변동가
                    try:
                        elem5 = driver.find_element_by_xpath('//*[@id="price"]/del').text
                        elem6 = driver.find_element_by_xpath('//*[@id="PricetoChange"]').text
                    except:
                        elem5 = driver.find_element_by_xpath('//*[@id="PricetoChange"]').text
                        elem6 = ' '

                    ws = wb.active

                    ws.cell(row=rowNum, column=1, value=elem1)  # 공급사 상품명
                    ws.cell(row=rowNum, column=2, value=elem2)  # 상품명
                    ws.cell(row=rowNum, column=3, value=elem3)  # 옵션입력
                    ws.cell(row=rowNum, column=4, value=elem4)  # 기타
                    ws.cell(row=rowNum, column=5, value=elem5)  # 공급가
                    ws.cell(row=rowNum, column=6, value=elem6)  # 변동가

                    wb.save("0" + idx + " " + text + "/#" + text + ".xlsx")

                    rowNum+=1

                    bs4 = BeautifulSoup(driver.page_source, "html.parser")
                    List4 = bs4.find('div', {'class': 'zoom-desc'}).findAll('a')

                    for n in range(0, len(List4)):
                        urlretrieve(List4[n]['href'],
                                    '01 BRANDS' + "/" + str(imgNum).rjust(5, '0') + "_img" + str(n + 1).rjust(2,
                                                                                                              '0') + ".jpg")
                    imgNum += 1

                    driver.back()

                try:
                    driver.find_element_by_xpath('//*[@title="Next page"]').click()
                except:
                    break

            driver.find_element_by_xpath('//*[@id="bs-example-navbar-collapse-1"]/ul/li[' + str(num) + ']/a').click()


def SALE(text, idx):
    title = ['10 Deep', '40s & Sies', '47', 'Acapulco Gold', 'Adidas', 'Akomplice', 'Almost August',
             'Alpha  Industries',
             'Asics', 'Baggu', 'Band Merch', 'Barney Cools', 'Baxter Of California', 'Billionaire Boys Club',
             'Black Scale',
             'Born X Raised', 'Brixton', 'Calvin Klein', 'Carhartt', 'Carrots', 'Champion', 'Cheap Monday', 'Chonies',
             'Clarks',
             'Clear Weather', 'CLSC', 'CMNDR', "Crooks & Castle Women's", 'Crooks & Castles', 'Diadora',
             'Diamond Supply Co.',
             'Dickies', 'Dimepiece', 'Dope', 'Elwood', 'Embellish', 'Ethika', 'Fairplay', 'Fourstar', 'Frank151',
             'Glassy Sunhaters', 'Goodworth', 'Gourmet', 'Grizzly Griptape', 'Herschel Supply', 'Herschel Supply Women',
             'Hex', 'Huf', 'I Love Ugly', ' in bloom', 'In4mation', 'Jansport', 'Jason Markk', 'Joyrich', 'Kidrobot',
             "Levi's Jeans", 'Love Me', 'LRG', 'Mitchell & Ness', 'Motel', 'Native', 'Nixon', 'Nixon Women',
             'Nudie Jeans',
             'OBEY', 'Odd Future', 'Onitsuka Tiger', 'P.F.Candle Co.', 'Pendleton', 'People Footwear',
             'Petals & Peacocks',
             'Primitive', 'Psychic Hearts', 'Publish', 'Puma', 'Raised By Wolves', 'Rastaclat', 'RIPNDIP', "Roi'al",
             'Rook',
             'Rothco', 'RVCA', 'Scotch & Soda', 'SSDD', 'SSUR', 'Stance', 'Standard Issue', 'Staple', 'Stussy',
             'Super Sunglasses',
             'Supra', 'The Hundreds', 'Thrasher', 'Timberland', 'Under Armour', 'Undftd', 'Unyforme', 'Uppercut Deluxe',
             'Valley Cruise Press', 'Vans', 'Vans Women', 'Veritas Aequitas', 'Visual', 'X - Large', 'Zanerobe']

    if not os.path.isdir("0" + idx + " " + text):
        os.mkdir("0" + idx + " " + text)

    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="공급사 상품명")  # 공급사 상품명
    ws.cell(row=1, column=2, value="상품명")  # 상품명
    ws.cell(row=1, column=3, value="옵션입력")  # 옵션입력
    ws.cell(row=1, column=4, value="기타")  # 기타
    ws.cell(row=1, column=5, value="공급가")  # 공급가
    ws.cell(row=1, column=6, value="변동가")  # 변동가

    wb.save("0" + idx + " " + text + "/#" + text + ".xlsx")

    bs4 = BeautifulSoup(driver.page_source, 'html.parser')
    ul = bs4.find('ul', id='moreBrand')
    List = ul.findAll('input')

    rowNum = 2
    imgNum = 1
    for i in range(0, len(List)):
        while True:
            answer = input(text + '-' + title[i] + "의 데이터를 추출하시겠습니까? y/n : ")
            if answer == "y":
                break
            elif answer == "n":
                break
            else:
                continue

        if answer == 'n':
            continue

        elif answer == 'y':
            if i != 0: driver.execute_script('RemoveFilterVal("' + List[i - 1]['id'] + '");')
            driver.execute_script(
                'document.getElementById("' + List[i]['id'] + '").setAttribute("checked", "checked");')
            driver.execute_script('javascript:GetResult();')

            try:
                driver.find_element_by_xpath('//*[@id="sortby"]/option[2]').click()
            except:
                driver.find_element_by_xpath('//*[@id="fysSort"]/option[3]').click()

            while True:
                bs4 = BeautifulSoup(driver.page_source, "html.parser")
                List = bs4.findAll('div', {'class': "item active"})

                for k in range(0, len(List)):
                    url = "https://www.mltd.com" + List[k].find('a')['href']
                    driver.get(url)

                    # 공급사 상품명
                    temp = url.partition('pr-')[2]
                    elem1 = temp.partition('.')[0]

                    # 상품명
                    a = driver.find_element_by_xpath('//*[@id="addprdct"]/p').text
                    elem2 = a.replace('\n', ' / ')

                    # 옵션입력, 기타
                    elem3 = ''
                    elem4 = ''
                    m = 1
                    while True:
                        try:
                            b = driver.find_element_by_xpath(
                                '//*[@id="addprdct"]/div[7]/table/tbody/tr[' + str(m) + ']/td[1]').text
                            if m == 1:
                                elem3 = b
                            else:
                                elem3 = elem3 + ' || ' + b
                            try:
                                btemp = b.partition(' / ')[0]
                            except:
                                btemp = b
                            c = driver.find_element_by_xpath(
                                '//*[@id="addprdct"]/div[7]/table/tbody/tr[' + str(m) + ']/td[2]').text
                            ctemp = c.split(' ')[0]
                            if ctemp == '':
                                ctemp = '100'
                            if m == 1:
                                elem4 += btemp + '-' + ctemp
                            else:
                                elem4 += '/' + btemp + '-' + ctemp
                            m += 1
                        except:
                            break

                    # 공급가 , 변동가
                    try:
                        elem5 = driver.find_element_by_xpath('//*[@id="price"]/del').text
                        elem6 = driver.find_element_by_xpath('//*[@id="PricetoChange"]').text
                    except:
                        elem5 = driver.find_element_by_xpath('//*[@id="PricetoChange"]').text
                        elem6 = ' '

                    ws = wb.active

                    ws.cell(row=rowNum, column=1, value=elem1)  # 공급사 상품명
                    ws.cell(row=rowNum, column=2, value=elem2)  # 상품명
                    ws.cell(row=rowNum, column=3, value=elem3)  # 옵션입력
                    ws.cell(row=rowNum, column=4, value=elem4)  # 기타
                    ws.cell(row=rowNum, column=5, value=elem5)  # 공급가
                    ws.cell(row=rowNum, column=6, value=elem6)  # 변동가

                    wb.save("0" + idx + " " + text + "/#" + text + ".xlsx")

                    rowNum+=1

                    bs4 = BeautifulSoup(driver.page_source, "html.parser")
                    List4 = bs4.find('div', {'class': 'zoom-desc'}).findAll('a')

                    for n in range(0, len(List4)):
                        urlretrieve(List4[n]['href'],
                                    '02 SALE' + "/" + str(imgNum).rjust(5, '0') + "_img" + str(n + 1).rjust(2,
                                                                                                              '0') + ".jpg")
                    imgNum += 1

                    driver.back()

                try:
                    driver.find_element_by_xpath('//*[@title="Next page"]').click()
                except:
                    break


if __name__ == "__main__":
    url = './chromedriver'  # 드라이브가 있는 경로
    driver = webdriver.Chrome(url)
    driver.set_window_size(1100, 1065)
    driver.get("http://www.mltd.com")

    P = 2
    for i in range(0, 4):
        if i == 2: P = 4
        data = driver.find_element_by_xpath('//*[@id="bs-example-navbar-collapse-1"]/ul/li[' + str(i + P) + ']/a')
        text = data.text

        while True:
            answer = input(text + "의 데이터를 추출하시겠습니까? y/n : ")
            if answer == "y":
                break
            elif answer == "n":
                break
            else:
                continue

        if answer == 'n':
            continue

        elif answer == 'y':
            data.click()

            if (i == 0) | (i == 2):
                BRANDS(text, i + P, str(i + 1))
            else:
                SALE(text, str(i + 1))